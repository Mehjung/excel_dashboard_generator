import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.formatting.rule import CellIsRule

# Beispieldaten erstellen
data = {
    'Datum': ['2024-01-01', '2024-01-01', '2024-02-01', '2024-02-01', '2024-03-01', '2024-03-01'],
    'Produkt': ['Produkt A', 'Produkt B', 'Produkt A', 'Produkt B', 'Produkt A', 'Produkt B'],
    'Region': ['Nord', 'Süd', 'Nord', 'Süd', 'Nord', 'Süd'],
    'Umsatz': [100, 200, 150, 250, 200, 300]
}

df = pd.DataFrame(data)

# Excel-Datei erstellen
file_path = 'professional_excel_dashboard.xlsx'
df.to_excel(file_path, index=False, sheet_name='Daten')

# Arbeitsmappe laden
wb = load_workbook(file_path)
ws = wb.active

# Pivot-Tabelle erstellen
pivot_data = pd.pivot_table(df, values='Umsatz', index=['Produkt'], columns=['Region'], aggfunc='sum', fill_value=0)
pivot_ws = wb.create_sheet(title='Pivot-Tabelle')

for r in dataframe_to_rows(pivot_data, index=True, header=True):
    pivot_ws.append(r)

# Diagramme mit Plotly erstellen

# Balkendiagramm
fig_bar = px.bar(df.groupby('Datum')['Umsatz'].sum().reset_index(), x='Datum', y='Umsatz', title='Umsatz nach Monat',
                 color_discrete_sequence=px.colors.qualitative.G10)
fig_bar.update_layout(title_font_size=20, title_font_family='Arial', title_font_color='darkblue')
fig_bar.write_image('bar_chart_plotly.png')

# Kreisdiagramm
fig_pie = px.pie(df, names='Produkt', values='Umsatz', title='Umsatzverteilung nach Produkt',
                 color_discrete_sequence=px.colors.qualitative.G10)
fig_pie.update_layout(title_font_size=20, title_font_family='Arial', title_font_color='darkblue')
fig_pie.write_image('pie_chart_plotly.png')

# Liniendiagramm
fig_line = px.line(df, x='Datum', y='Umsatz', color='Produkt', markers=True, title='Umsatzentwicklung nach Produkt',
                   color_discrete_sequence=px.colors.qualitative.G10)
fig_line.update_layout(title_font_size=20, title_font_family='Arial', title_font_color='darkblue')
fig_line.write_image('line_chart_plotly.png')

# Diagramme in Excel einfügen
charts = [
    {'image': 'bar_chart_plotly.png', 'anchor': 'B10'},
    {'image': 'pie_chart_plotly.png', 'anchor': 'B30'},
    {'image': 'line_chart_plotly.png', 'anchor': 'B50'}
]

for chart in charts:
    img = Image(chart['image'])
    img.anchor = chart['anchor']
    ws.add_image(img)

# Formatierungen und Stile anwenden
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=14)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
cell_alignment = Alignment(horizontal='center', vertical='center')

for cell in ws['1:1']:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = cell_alignment

# Zusätzliche Formatierung für Pivot-Tabelle
pivot_ws['A1'].font = Font(bold=True)
pivot_ws['B1'].font = Font(bold=True)
pivot_ws['C1'].font = Font(bold=True)
for cell in pivot_ws['A:A']:
    cell.font = Font(bold=True)
pivot_ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
pivot_ws['B1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
pivot_ws['C1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Bedingte Formatierung hinzufügen
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
pivot_ws.conditional_formatting.add('B2:C10', CellIsRule(operator='lessThan', formula=['150'], fill=red_fill))
pivot_ws.conditional_formatting.add('B2:C10', CellIsRule(operator='greaterThanOrEqual', formula=['150'], fill=green_fill))

# Arbeitsmappe speichern
wb.save(file_path)

print(f'Professionelles Fluent Design Dashboard erstellt und gespeichert in {file_path}')
